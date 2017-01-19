import openpyxl

industries = []

wb = openpyxl.load_workbook('CompanyInformation.xlsx')
ws = wb.get_sheet_by_name('CompanyInformation')
for row in ws.iter_rows('C2:C45'.format(ws.min_row, ws.max_row)):
    for cell in row:
        industries.append(cell.value)

i = 0
while i < len(industries):
    industries[i] = industries[i].encode("utf-8")
    i = i + 1

print industries
print len(industries)        
