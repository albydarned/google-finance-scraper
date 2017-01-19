import openpyxl

tickers = []

wb = openpyxl.load_workbook('CompanyInformation.xlsx')
ws = wb.get_sheet_by_name('CompanyInformation')
for row in ws.iter_rows('A2:A46'.format(ws.min_row, ws.max_row)):
    for cell in row:
        tickers.append(cell.value)

i = 0
while i < len(tickers):
    tickers[i] = tickers[i].encode("utf-8")
    i = i + 1

print tickers
print len(tickers)        
